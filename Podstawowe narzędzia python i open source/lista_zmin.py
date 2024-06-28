import pandas as pd
import openpyxl
import os
import requests
import re

def download_file(url, folder):
    # Pobierz plik z podanego adresu URL
    response = requests.get(url)
    # Wyodrębnij nazwę pliku z adresu URL
    file_name = url.split("/")[-1]
    # Usuń nieprawidłowe znaki z nazwy pliku
    file_name = re.sub(r'[\\/*?:"<>|]', '', file_name)
    # Utwórz ścieżkę docelową
    file_path = os.path.join(folder, file_name)
    # Zapisz plik na dysku
    with open(file_path, "wb") as file:
        file.write(response.content)
    print(f"Pobrano plik: {file_name}")
    return file_path




def normalize_data(file_path):
    # Wczytaj plik Excel
    workbook = openpyxl.load_workbook(file_path)
    # Dla każdego arkusza
    for sheet in workbook.sheetnames:
        # Wybierz arkusz
        ws = workbook[sheet]
        # Dla każdej komórki w arkuszu
        for row in ws.iter_rows(values_only=True):
            # Normalizuj dane w komórce
            normalized_row = [str(cell).strip().capitalize() if isinstance(cell, str) else cell for cell in row]
            # Ustaw znormalizowane dane w komórce
            for i, value in enumerate(normalized_row):
                ws.cell(row=row[0], column=i + 1, value=value)
    # Zapisz zmiany w pliku
    workbook.save(file_path)
    print(f"Znormalizowano dane w pliku: {file_path}")


def compare_data(file_path_old, file_path_new):
    # Wczytaj dane z obu plików Excel
    df_old = pd.read_excel(file_path_old)
    df_new = pd.read_excel(file_path_new)

    # Jeśli kolumna 'ID' nie istnieje, użyj pierwszej kolumny jako indeks
    if 'ID' not in df_old.columns:
        df_old.set_index(df_old.columns[0], inplace=True)
    if 'ID' not in df_new.columns:
        df_new.set_index(df_new.columns[0], inplace=True)

    # Znajdź nowe rekordy (w nowym pliku, ale nie w starym)
    new_records = df_new.index.difference(df_old.index)

    # Znajdź usunięte rekordy (w starym pliku, ale nie w nowym)
    deleted_records = df_old.index.difference(df_new.index)

    # Znajdź zmienione rekordy (obecne w obu plikach)
    changed_records = df_old.index.intersection(df_new.index)

    # Stwórz raport
    report = []

    # Dodaj nowe rekordy do raportu
    for index in new_records:
        record = {'ID': index, 'Status': 'Nowy', 'Zmiany': 'Brak zmian'}
        report.append(record)

    # Dodaj usunięte rekordy do raportu
    for index in deleted_records:
        record = {'ID': index, 'Status': 'Usunięty', 'Zmiany': 'Brak zmian'}
        report.append(record)

    # Porównaj zmienione rekordy i dodaj informacje o zmianach do raportu
    for index in changed_records:
        changes = []
        for column in df_old.columns:
            if df_old.loc[index, column] != df_new.loc[index, column]:
                change = f"{column}: '{df_old.loc[index, column]}' -> '{df_new.loc[index, column]}'"
                changes.append(change)
        if changes:
            record = {'ID': index, 'Status': 'Zmieniony', 'Zmiany': ', '.join(changes)}
            report.append(record)

    # Zwróć raport jako DataFrame
    report_df = pd.DataFrame(report)

    return report_df


def generate_change_list(report_df, category):
    new_records = report_df[report_df['Status'] == 'Nowy']
    deleted_records = report_df[report_df['Status'] == 'Usunięty']
    changed_records = report_df[report_df['Status'] == 'Zmieniony']

    change_list = []

    if not new_records.empty:
        new_ids = ', '.join(new_records['ID'].tolist())
        change_list.append(f"Nowe {category}: {new_ids}")

    if not deleted_records.empty:
        deleted_ids = ', '.join(deleted_records['ID'].tolist())
        change_list.append(f"{category} wycofane z oferty: {deleted_ids}")

    if not changed_records.empty:
        for column in changed_records['Zmiany']:
            column_name = column.split(":")[0].strip()
            changed_ids = ', '.join(changed_records['ID'].tolist())
            change_list.append(f"Kolumna {column_name} zmieniła się dla {category.lower()} : {changed_ids}")

    return change_list


def main():
    # Ścieżki do plików .xlsx
    file_path_old = "download/katalog_2023.xlsx"
    file_path_new = "download/katalog_2024.xlsx"

    # Pobierz i znormalizuj pliki
    normalize_data(
        download_file("https://drive.google.com/uc?export=download&id=1wKpSpTx89dbU3SrKt-3-DqQ62kHOFHSX", "download"))
    normalize_data(
        download_file("https://drive.google.com/uc?export=download&id=1oYKyW7flL53smo56W9srpFdoUuz6_42x", "download"))

    # Porównaj dane
    report = compare_data(file_path_old, file_path_new)

    # Zapisz raport do pliku CSV
    report.to_csv("comparison_report.csv", index=False)
    print("Porównanie zakończone. Raport został zapisany do pliku comparison_report.csv")

    # Generuj listę zmian
    report_df = pd.read_csv("comparison_report.csv")
    elektronarzedzia_changes = generate_change_list(report_df[report_df['ID'].str.startswith('E')], "Elektronarzędzia")
    ostrza_changes = generate_change_list(report_df[report_df['ID'].str.startswith('O')], "Ostrza")
    change_list = elektronarzedzia_changes + ostrza_changes
    print("\n".join(change_list))


if __name__ == "__main__":
    main()
