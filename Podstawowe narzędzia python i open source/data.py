import openpyxl


def normalize_data(file_path):
    # Wczytaj plik Excel
    workbook = openpyxl.load_workbook(file_path)
    # Wybierz pierwszy arkusz
    sheet = workbook.active

    # Sortowanie danych w kolumnie Zastosowanie w arkuszu Ostrza
    if sheet.title == "Ostrza":
        zastosowanie_column = sheet["D"]
        sorted_zastosowanie_column = sorted(zastosowanie_column, key=lambda x: x.value)
        for i, cell in enumerate(zastosowanie_column, start=1):
            cell.value = sorted_zastosowanie_column[i - 1].value

    # Kapitalizacja danych w arkuszu Elektronarzędzia
    if sheet.title == "Elektronarzędzia":
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=4):
            new_row = [str(cell.value).strip().capitalize() if isinstance(cell.value, str) else cell.value for cell in
                       row]
            for i, cell in enumerate(row):
                cell.value = new_row[i]

    # Kapitalizacja danych w arkuszu Ostrza
    if sheet.title == "Ostrza":
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=5):
            new_row = [str(cell.value).strip().capitalize() if isinstance(cell.value, str) else cell.value for cell in
                       row]
            for i, cell in enumerate(row):
                cell.value = new_row[i]

    # Wyczyść nadmiarowe białe znaki
    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column, values_only=True):
        new_row = [value.strip() if isinstance(value, str) else value for value in row]
        # Utwórz licznik wierszy
        row_number = 1
        for i, value in enumerate(new_row):
            # Ustaw wartość komórki na podstawie licznika wierszy
            sheet.cell(row=row_number, column=i + 1, value=value)
        # Inkrementuj licznik wierszy
        row_number += 1

    # Zapisz zmiany
    workbook.save(file_path)
    print(f"Znormalizowano dane w pliku: {file_path}")


def main():
    # Lista ścieżek do plików .xlsx
    files = ["download/katalog_2023.xlsx", "download/katalog_2024.xlsx"]

    # Normalizacja danych dla każdego pliku
    for file_path in files:
        normalize_data(file_path)


if __name__ == "__main__":
    main()
